"""
Patent Scraper — SerpAPI Google Patents
Produces one Excel file with two sheets:
  1. Granted patents only
  2. All activity (grants + applications)
Plus a summary txt file.

Free tier: 250 searches/month at serpapi.com
Install: pip install requests pandas openpyxl pyyaml beautifulsoup4 matplotlib scikit-learn
"""

import subprocess
import sys
import os
import time
import json
import threading
import sys as _sys
from concurrent.futures import ThreadPoolExecutor, as_completed

# Python version check
if sys.version_info < (3, 8):
    sys.exit(
        f"ERROR: Python 3.8 or higher is required. "
        f"You are running Python {sys.version_info.major}.{sys.version_info.minor}. "
        f"Download the latest version at https://www.python.org/downloads/"
    )

# Auto-install missing dependencies
def install_requirements():
    pkg_map = {
        "requests":       "requests",
        "pandas":         "pandas",
        "openpyxl":       "openpyxl",
        "pyyaml":         "yaml",
        "beautifulsoup4": "bs4",
        "matplotlib":     "matplotlib",
        "scikit-learn":   "sklearn",
        "wordcloud":      "wordcloud",
    }
    import importlib.util
    missing = [pkg for pkg, imp in pkg_map.items() if not importlib.util.find_spec(imp)]

    if not missing:
        print("All packages already installed.\n")
        return

    print("The following packages are missing and will be installed:")
    for pkg in missing:
        try:
            result = subprocess.run(
                [sys.executable, "-m", "pip", "index", "versions", pkg],
                capture_output=True, text=True
            )
            version = "unknown"
            for line in result.stdout.splitlines():
                if "LATEST" in line:
                    version = line.split(":")[-1].strip().split(",")[0].strip()
                    break
            print(f"  {pkg} (latest: {version})")
        except Exception:
            print(f"  {pkg}")

    subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])
    print("Done.\n")

if input("Auto-install missing packages if needed? (y/n): ").strip().lower() != "n":
    install_requirements()

import yaml
import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from datetime import datetime

# Load API key from config.yaml
with open("config.yaml") as f:
    config = yaml.safe_load(f)
API_KEY       = config.get("api_key", "")
MONTHLY_LIMIT = config.get("monthly_limit", 250)
if not API_KEY or API_KEY == "your_serpapi_key_here":
    raise SystemExit("ERROR: Set your SerpAPI key in config.yaml (copy from config.yaml.example)")

ASSIGNEE = input("Enter assignee name (e.g. Thomas Jefferson University): ").strip()

# Date range or single start date
print("\nDate filter options:")
print("  1. From a start date through present (e.g. 20250101)")
print("  2. Date range (e.g. 20230101 to 20251231)")
date_choice = input("Choose option (1 or 2): ").strip()

if date_choice == "2":
    AFTER_DATE  = input("Enter start date (YYYYMMDD): ").strip()
    BEFORE_DATE = input("Enter end date (YYYYMMDD): ").strip()
    DATE_LABEL  = f"{AFTER_DATE}_to_{BEFORE_DATE}"
else:
    AFTER_DATE  = input("Enter start date (YYYYMMDD): ").strip()
    BEFORE_DATE = None
    DATE_LABEL  = AFTER_DATE

PARALLEL = input("Fetch inventor/assignee data in parallel? (y/n, faster but may get rate-limited): ").strip().lower() == "y"
if PARALLEL:
    try:
        WORKERS = int(input("How many parallel workers? (recommended: 3-10, default 5): ").strip() or "5")
    except ValueError:
        WORKERS = 5
else:
    WORKERS = 5

TIMESTAMP  = datetime.now().strftime('%Y%m%d_%H%M%S')
SHORT_TAG  = "".join(w[0].upper() for w in ASSIGNEE.split())
OUTPUT_DIR = f"search_{TIMESTAMP}_{ASSIGNEE.replace(' ', '_')}_{DATE_LABEL}"
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT     = os.path.join(OUTPUT_DIR, f"{SHORT_TAG}_{DATE_LABEL}_patents.xlsx")
ENDPOINT   = "https://serpapi.com/search"
CACHE_FILE = "patent_cache.json"

api_credits_used = 0


# ---------------------------------------------------------------------------
# Cache
# ---------------------------------------------------------------------------
def load_cache():
    try:
        with open(CACHE_FILE) as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_cache(cache):
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2)


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Monthly usage tracker
# ---------------------------------------------------------------------------
USAGE_FILE = "usage_tracker.json"

def load_usage():
    current_month = datetime.now().strftime("%Y-%m")
    try:
        with open(USAGE_FILE) as f:
            data = json.load(f)
        if data.get("month") != current_month:
            data = {"month": current_month, "credits_used": 0}
    except FileNotFoundError:
        data = {"month": current_month, "credits_used": 0}
    return data

def save_usage(data):
    with open(USAGE_FILE, "w") as f:
        json.dump(data, f, indent=2)

def display_usage(credits_this_run):
    usage = load_usage()
    usage["credits_used"] += credits_this_run
    save_usage(usage)
    remaining = max(0, MONTHLY_LIMIT - usage["credits_used"])
    print(f"\nAPI credits used this run:    {credits_this_run}")
    print(f"API credits used this month:  {usage['credits_used']} / {MONTHLY_LIMIT}")
    print(f"Credits remaining this month: {remaining}")
    if remaining < 20:
        print(f"  \u26a0\ufe0f  Warning: only {remaining} credits remaining this month.")
    print(f"Check live balance at: https://serpapi.com/dashboard")


# Scrape inventors + co-assignees from patent page
# ---------------------------------------------------------------------------
def fetch_patent_details(patent_number, patent_link, serpapi_assignee):
    try:
        resp = requests.get(patent_link, timeout=15, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        })
        soup = BeautifulSoup(resp.text, "html.parser")

        inventors = [tag.get_text(strip=True) for tag in soup.find_all(itemprop="inventor")]
        inv_str   = ", ".join(inventors) if inventors else "N/A"

        assignees = []
        for dt in soup.find_all("dt"):
            if "current assignee" in dt.get_text(strip=True).lower():
                for sibling in dt.find_next_siblings():
                    if sibling.name == "dt":
                        break
                    if sibling.name == "dd":
                        text = sibling.get_text(strip=True)
                        if text:
                            assignees.append(text)
                break

        def is_main(scraped, ref):
            s, r = scraped.lower(), ref.lower()
            return s == r or s in r or r in s

        co_assignees = [
            a for a in assignees
            if not is_main(a, serpapi_assignee) and not is_main(a, ASSIGNEE)
        ]
        co_str = ", ".join(co_assignees) if co_assignees else "None"

        return inv_str, co_str
    except Exception as e:
        print(f"  ERROR fetching {patent_number}: {e}")
        return "N/A", "None"


# ---------------------------------------------------------------------------
# Enrich patents
# ---------------------------------------------------------------------------
def enrich_patents(patents, label=""):
    cache   = load_cache()
    changed = False
    print(f"\nEnriching {label} ({len(patents)} patents) — parallel={PARALLEL}...")

    uncached = [p for p in patents if p["Patent Number"] not in cache]
    cached   = [p for p in patents if p["Patent Number"] in cache]

    for p in cached:
        entry = cache[p["Patent Number"]]
        p["Inventors"]    = entry["inventors"]
        p["Co-Assignees"] = entry["co_assignees"]
        print(f"  {p['Patent Number']}: {entry['inventors']} | co: {entry['co_assignees']} (cached)")

    if uncached:
        if PARALLEL:
            with ThreadPoolExecutor(max_workers=WORKERS) as executor:
                futures = {
                    executor.submit(
                        fetch_patent_details,
                        p["Patent Number"], p["Link"], p.get("Assignee", ASSIGNEE)
                    ): p for p in uncached
                }
                for i, future in enumerate(as_completed(futures)):
                    p = futures[future]
                    inv_str, co_str = future.result()
                    p["Inventors"]    = inv_str
                    p["Co-Assignees"] = co_str
                    cache[p["Patent Number"]] = {"inventors": inv_str, "co_assignees": co_str}
                    changed = True
                    print(f"  [{i+1}/{len(uncached)}] {p['Patent Number']}: {inv_str} | co: {co_str}")
        else:
            for i, p in enumerate(uncached):
                inv_str, co_str = fetch_patent_details(
                    p["Patent Number"], p["Link"], p.get("Assignee", ASSIGNEE)
                )
                p["Inventors"]    = inv_str
                p["Co-Assignees"] = co_str
                cache[p["Patent Number"]] = {"inventors": inv_str, "co_assignees": co_str}
                changed = True
                print(f"  [{i+1}/{len(uncached)}] {p['Patent Number']}: {inv_str} | co: {co_str}")
                time.sleep(0.5)

    if changed:
        save_cache(cache)
        print(f"  Cache updated → {CACHE_FILE}")

    return patents


# ---------------------------------------------------------------------------
# SerpAPI fetch
# ---------------------------------------------------------------------------
def fetch_page(page, status=None):
    global api_credits_used
    params = {
        "engine":   "google_patents",
        "q":        "",
        "assignee": ASSIGNEE,
        "after":    f"publication:{AFTER_DATE}",
        "num":      100,
        "page":     page,
        "api_key":  API_KEY,
    }
    if BEFORE_DATE:
        params["before"] = f"publication:{BEFORE_DATE}"
    if status:
        params["status"] = status
    resp = requests.get(ENDPOINT, params=params, timeout=30)
    resp.raise_for_status()
    api_credits_used += 1
    return resp.json()


def fetch_all_pages(status=None, label=""):
    patents = []
    page    = 1
    while True:
        print(f"  [{label}] Page {page}...", end=" ", flush=True)
        try:
            data = fetch_page(page, status=status)
        except Exception as e:
            print(f"Error: {e}")
            break
        if page == 1:
            info = data.get("search_information", {})
            print(f"\n  → {info.get('total_results', '?')} total results")
        results = data.get("organic_results", [])
        if not results:
            print("No results — done.")
            break
        for r in results:
            patents.append({
                "Title":            r.get("title", "N/A"),
                "Patent Number":    r.get("publication_number", "N/A"),
                "Publication Date": r.get("publication_date", "N/A"),
                "Grant Date":       r.get("grant_date", "N/A"),
                "Filing Date":      r.get("filing_date", "N/A"),
                "Priority Date":    r.get("priority_date", "N/A"),
                "Inventors":        r.get("inventor", "N/A"),
                "Assignee":         r.get("assignee", "N/A"),
                "Co-Assignees":     "None",  # always string, never NaN
                "Abstract":         r.get("snippet", "N/A"),
                "Link":             r.get("patent_link", "N/A"),
            })
        print(f"got {len(results)} (total: {len(patents)})")
        if not data.get("serpapi_pagination", {}).get("next"):
            print("  Done.")
            break
        page += 1
        time.sleep(1)
    return patents


# ---------------------------------------------------------------------------
# Assignee review
# ---------------------------------------------------------------------------
def assignee_review(granted, all_patents):
    all_pats = granted + [
        p for p in all_patents
        if p["Patent Number"] not in {x["Patent Number"] for x in granted}
    ]
    unique_assignees = sorted(set(
        p.get("Assignee", "").strip()
        for p in all_pats
        if p.get("Assignee", "").strip()
    ))

    print("\n" + "=" * 60)
    print(f"ASSIGNEE REVIEW — the following organizations were returned")
    print(f"by SerpAPI for your search: '{ASSIGNEE}'")
    print("Some may not be the institution you are looking for.")
    print("Enter the numbers to EXCLUDE (comma-separated), or press")
    print("Enter to keep all:\n")
    for i, a in enumerate(unique_assignees, 1):
        print(f"  {i}. {a}")

    # Countdown timer — auto-proceeds with all assignees after 30s if unattended
    TIMEOUT = 30
    user_input = []
    input_event = threading.Event()

    def get_input():
        try:
            val = input("\nExclude numbers (or Enter to keep all): ").strip()
            user_input.append(val)
        except Exception:
            user_input.append("")
        input_event.set()

    input_thread = threading.Thread(target=get_input, daemon=True)
    input_thread.start()

    for remaining in range(TIMEOUT, 0, -1):
        if input_event.is_set():
            break
        print(f"\r  Auto-proceeding with all assignees in {remaining}s... (type to cancel)  ", end="", flush=True)
        time.sleep(1)

    if not input_event.is_set():
        print(f"\r  No response — proceeding with all assignees.                              ")
        raw = ""
    else:
        print()
        raw = user_input[0] if user_input else ""

    excluded = set()
    excluded_patents = {}

    if raw:
        try:
            for n in raw.split(","):
                idx = int(n.strip()) - 1
                if 0 <= idx < len(unique_assignees):
                    name = unique_assignees[idx]
                    excluded.add(name)
                    excluded_patents[name] = [
                        p for p in all_pats if p.get("Assignee", "") == name
                    ]
        except ValueError:
            print("  Invalid input — keeping all.")

    if excluded:
        print(f"  Excluding: {excluded}")
        granted     = [p for p in granted     if p.get("Assignee", "") not in excluded]
        all_patents = [p for p in all_patents if p.get("Assignee", "") not in excluded]

    print(f"  Final: {len(granted)} granted, {len(all_patents)} all activity")
    print("=" * 60)
    return granted, all_patents, excluded_patents


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------
def prepare_df(patents):
    df = pd.DataFrame(patents)
    df = df.groupby("Patent Number", as_index=False).agg({
        "Title": lambda x: " / ".join(x.unique()),
        **{col: "first" for col in df.columns if col not in ["Patent Number", "Title"]}
    })
    df = df.sort_values("Publication Date", ascending=True)
    # Normalize Co-Assignees — ensure NaN becomes string "None"
    if "Co-Assignees" in df.columns:
        df["Co-Assignees"] = df["Co-Assignees"].fillna("None").replace("", "None")
    cols = ["Title", "Patent Number", "Publication Date", "Grant Date",
            "Filing Date", "Priority Date", "Inventors", "Assignee",
            "Co-Assignees", "Abstract", "Link"]
    df = df[[c for c in cols if c in df.columns]]
    return df


def save_excel(df_granted, df_all, path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_granted.to_excel(writer, sheet_name="Granted", index=False)
        df_all.to_excel(writer, sheet_name="All Activity", index=False)
    wb = openpyxl.load_workbook(path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.font = Font(bold=True)
    wb.save(path)
    print(f"\n✓ Saved to {path}")
    print(f"  Sheet 'Granted': {len(df_granted)} patents")
    print(f"  Sheet 'All Activity': {len(df_all)} patents")


def save_summary(df_granted, df_all, granted_raw, all_raw, excluded_patents, path):
    originally_missing = {
        p["Patent Number"] for p in granted_raw
        if p["Patent Number"] not in {x["Patent Number"] for x in all_raw}
    }
    granted_dupes = df_granted[df_granted["Title"].str.contains(" / ", na=False)]
    all_dupes     = df_all[df_all["Title"].str.contains(" / ", na=False)]

    with open(path, "w") as f:
        f.write(f"Patent Search Summary\n")
        f.write(f"Assignee: {ASSIGNEE}\n")
        f.write(f"Date filter: from {AFTER_DATE}" + (f" to {BEFORE_DATE}" if BEFORE_DATE else " through present") + "\n")
        f.write(f"Generated: {TIMESTAMP}\n")
        f.write("=" * 60 + "\n\n")

        f.write(f"GRANTED ({len(df_granted)} patents):\n")
        for _, row in df_granted.iterrows():
            f.write(f"  {row['Patent Number']} | {row.get('Grant Date','N/A')} | {row['Title'][:80]}\n")

        f.write(f"\nALL ACTIVITY ({len(df_all)} patents):\n")
        for _, row in df_all.iterrows():
            status = "Granted" if pd.notna(row.get("Grant Date")) and str(row.get("Grant Date")) not in ("NaT", "nan", "N/A", "") else "Application"
            # Normalize co-assignees — replace NaN with string "None"
            f.write(f"  [{status}] {row['Patent Number']} | {row.get('Publication Date','N/A')} | {row['Title'][:80]}\n")

        f.write(f"\nDUPLICATE TITLES IN GRANTED ({len(granted_dupes)}):\n")
        for _, row in granted_dupes.iterrows():
            f.write(f"  {row['Patent Number']} | {row['Title']}\n")
        if len(granted_dupes) == 0:
            f.write("  None\n")

        f.write(f"\nDUPLICATE TITLES IN ALL ACTIVITY ({len(all_dupes)}):\n")
        for _, row in all_dupes.iterrows():
            f.write(f"  {row['Patent Number']} | {row['Title']}\n")
        if len(all_dupes) == 0:
            f.write("  None\n")

        f.write(f"\nADDED TO ALL ACTIVITY FROM GRANTED ({len(originally_missing)}):\n")
        if originally_missing:
            for num in sorted(originally_missing):
                row = df_granted[df_granted["Patent Number"] == num].iloc[0]
                f.write(f"  {num} | {row.get('Grant Date','N/A')} | {row['Title'][:80]}\n")
        else:
            f.write("  None\n")

        # Excluded institutions and their patents
        f.write(f"\nEXCLUDED INSTITUTIONS ({len(excluded_patents)}):\n")
        if excluded_patents:
            for institution, pats in sorted(excluded_patents.items()):
                f.write(f"\n  {institution} ({len(pats)} patents):\n")
                for p in pats:
                    f.write(f"    {p['Patent Number']} | {p['Title'][:70]}\n")
        else:
            f.write("  None\n")

    print(f"✓ Summary written to {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def scrape():
    date_desc = f"from {AFTER_DATE}" + (f" to {BEFORE_DATE}" if BEFORE_DATE else " through present")
    print(f"\n=== SerpAPI Google Patents ===")
    print(f"Assignee: '{ASSIGNEE}', {date_desc}, parallel={PARALLEL}\n")

    print("--- Query 1: Granted patents ---")
    granted = fetch_all_pages(status="GRANT", label="GRANT")
    granted = enrich_patents(granted, label="Granted")

    print("\n--- Query 2: All activity ---")
    all_patents = fetch_all_pages(status=None, label="ALL")
    all_patents = enrich_patents(all_patents, label="All Activity")

    granted, all_patents, excluded_patents = assignee_review(granted, all_patents)

    if granted or all_patents:
        df_granted = prepare_df(granted) if granted else pd.DataFrame()

        merged = all_patents.copy()
        all_numbers = {p["Patent Number"] for p in merged}
        for p in granted:
            if p["Patent Number"] not in all_numbers:
                merged.append(p)

        df_all = prepare_df(merged) if merged else pd.DataFrame()

        save_excel(df_granted, df_all, OUTPUT)
        summary_path = os.path.join(OUTPUT_DIR, f"{SHORT_TAG}_{DATE_LABEL}_summary.txt")
        save_summary(df_granted, df_all, granted, all_patents, excluded_patents, summary_path)

    display_usage(api_credits_used)


if __name__ == "__main__":
    scrape()
